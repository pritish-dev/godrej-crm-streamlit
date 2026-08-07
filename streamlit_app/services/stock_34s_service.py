"""
services/stock_34s_service.py

34S Physical Stock Register — horizontal pivot, one Google Sheet tab per month.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SHEET DESIGN
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Tab name : "34s Stock Register- May 2026"
           "34s Stock Register- Jun 2026"   (auto-created each month)

Layout (grows RIGHT by 5 columns each day):

  Sl No │ Item Code │ Item Description │ Product Category │ 01/05 Op Stock │ 01/05 In Ward │ 01/05 Out Ward │ 01/05 Cl Stock │ 01/05 DC No │ 02/05 Op Stock │ …

  • 1 row per item — never changes
  • 5 columns per day — appended each evening at 8 PM
  • ~1,000 items × (4 + 5 × 31) = ~159 K cells/month  ✅

DAILY JOB (8 PM IST):
  1. Ensure month tab exists (auto-copies item list from previous month if new)
  2. _get_prev_cl_stock()  →  most recent previous day's Cl Stock (crosses months)
  3. Fetch In Ward  (email "Delivery Challan Information" + Drive PDFs, ZBF34S only)
  4. Fetch Out Ward (34S PHYSICAL DELIVERY CHALLAN + 34S RETURN RPL sheets)
  5. Build 5 new columns, drop existing ones for that date (idempotent)
  6. Write the combined DataFrame back to the tab

CATCH-UP:
  run_update_range(start, end) updates every missing date in sequence,
  clearing the gspread cache between days so each day reads fresh data.
"""

from __future__ import annotations

import io
import os
import re
import sys
import gspread
from calendar import monthrange
from datetime import date, datetime, timedelta, timezone

import pandas as pd

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, BASE_DIR)

from utils.helpers import to_indian_number_string

# ─── Constants ────────────────────────────────────────────────────────────────

IST             = timezone(timedelta(hours=5, minutes=30))
SHEET_PREFIX    = "34s Stock Register- "          # + "May 2026"
FIXED_COLS      = ["Sl No", "Item Code", "Item Description", "Product Category"]
DATE_SUB_COLS   = ["Op Stock", "In Ward", "Out Ward", "Cl Stock", "DC No"]
WAREHOUSE_CODE  = "ZBF34S"
CHALLAN_SUBJECT = "Delivery Challan Information"
DELIVERY_SHEET  = "34S PHYSICAL DELIVERY CHALLAN"
RETURN_SHEET    = "34S RETURN RPL"

# ─── Flat single-day register format ──────────────────────────────────────────
# The live OPS tab ("34s Stock Register- August 2026") is a *flat* register, not
# the horizontal date-pivot this module was originally written for:
#
#   Row 1 : title banner            "PHYSICAL STOCK REGISTER 34S … 07 August 2026"
#   Row 2 : column headers          Sl No | Item code | Item Description |
#                                    Product Category | Op Stock | In ward |
#                                    Out ward | Cl Stock
#   Row 3+: one row per item, a single Op/In/Out/Cl block (no date prefixes,
#           no DC No column), overwritten in place each day.
#
# The four movement columns are matched by header label; the *item code* used to
# match inward/outward is detected by CONTENT (the column whose cells match the
# Godrej item-code pattern), because on the live sheet the "Item code" header
# actually sits over the warehouse code (ZBF34S) while the real code lives under
# the "Item Description" header.
FLAT_MOVE_COLS  = ["Op Stock", "In Ward", "Out Ward", "Cl Stock"]
_ITEM_CODE_RE   = re.compile(r"\d{8}[A-Z]{2}\d{5}")

# Header-variant → canonical name, for BOTH the fixed and the movement columns
# of the flat register.  Compared case-insensitively with whitespace collapsed.
_FLAT_SYNONYMS: dict[str, str] = {
    # fixed
    "sl no": "Sl No", "sl. no": "Sl No", "sl.no": "Sl No", "slno": "Sl No",
    "s no": "Sl No", "sr no": "Sl No", "serial no": "Sl No", "sno": "Sl No",
    "item code": "Item Code", "itemcode": "Item Code", "item no": "Item Code",
    "material code": "Item Code", "product code": "Item Code", "sku": "Item Code",
    "item description": "Item Description", "description": "Item Description",
    "item desc": "Item Description", "item name": "Item Description",
    "product category": "Product Category", "category": "Product Category",
    "prod category": "Product Category", "item category": "Product Category",
    # movement
    "op stock": "Op Stock", "opening stock": "Op Stock", "opening": "Op Stock",
    "in ward": "In Ward", "inward": "In Ward", "in-ward": "In Ward",
    "in word": "In Ward", "received": "In Ward",
    "out ward": "Out Ward", "outward": "Out Ward", "out-ward": "Out Ward",
    "out word": "Out Ward", "issued": "Out Ward", "dispatched": "Out Ward",
    "cl stock": "Cl Stock", "closing stock": "Cl Stock", "closing": "Cl Stock",
}

# ─── Sheet naming ─────────────────────────────────────────────────────────────

def sheet_name_for(d: date) -> str:
    """e.g. '34s Stock Register- May 2026'"""
    return f"{SHEET_PREFIX}{d.strftime('%B %Y')}"


def current_sheet_name() -> str:
    return sheet_name_for(datetime.now(IST).date())


# ─── Column naming helpers ────────────────────────────────────────────────────

def _col_tag(d: date) -> str:
    """Short date prefix for column names: '21/05'"""
    return f"{d.day:02d}/{d.month:02d}"


def _col(d: date, sub: str) -> str:
    """Full column name: '21/05 Op Stock'"""
    return f"{_col_tag(d)} {sub}"


_COL_RE = re.compile(
    r"^(\d{2})/(\d{2})\s+(Op Stock|In Ward|Out Ward|Cl Stock|DC No)$"
)


def _parse_col(col_name: str) -> tuple[int, int, str] | None:
    """
    Parse '21/05 Op Stock' → (day=21, month=5, sub='Op Stock').
    Returns None if the name doesn't match.
    """
    m = _COL_RE.match(col_name.strip())
    if not m:
        return None
    return int(m.group(1)), int(m.group(2)), m.group(3)


def dates_in_df(df: pd.DataFrame, year: int, month: int) -> list[date]:
    """Return sorted list of dates that have at least one column group in df."""
    seen: set[date] = set()
    for col in df.columns:
        parsed = _parse_col(col)
        if not parsed:
            continue
        day, mon, _ = parsed
        if mon != month:
            continue
        try:
            seen.add(date(year, month, day))
        except ValueError:
            pass
    return sorted(seen)


# ─── Fixed-column header normalization ────────────────────────────────────────
# The sheet is often created / edited by hand, so the four fixed columns show up
# with inconsistent casing, spacing or wording ("ITEM CODE", "Sl. No",
# "Category", …).  Downstream logic matches these columns by their exact
# canonical name, so a mismatch makes the whole sheet look empty
# ("No items found …").  We map common variants back to the canonical name on
# read.  Keys are compared case-insensitively with internal whitespace collapsed.

_FIXED_COL_SYNONYMS: dict[str, str] = {
    # → Sl No
    "sl no": "Sl No", "sl. no": "Sl No", "sl.no": "Sl No", "slno": "Sl No",
    "sl no.": "Sl No", "s no": "Sl No", "s. no": "Sl No", "s.no": "Sl No",
    "sr no": "Sl No", "sr. no": "Sl No", "sr.no": "Sl No",
    "serial no": "Sl No", "serial number": "Sl No", "sno": "Sl No",
    # → Item Code
    "item code": "Item Code", "itemcode": "Item Code", "item no": "Item Code",
    "item number": "Item Code", "material code": "Item Code",
    "material no": "Item Code", "material number": "Item Code",
    "product code": "Item Code", "sku": "Item Code", "sku code": "Item Code",
    # → Item Description
    "item description": "Item Description", "item desc": "Item Description",
    "description": "Item Description", "desc": "Item Description",
    "item name": "Item Description", "material description": "Item Description",
    "product description": "Item Description", "product name": "Item Description",
    # → Product Category
    "product category": "Product Category", "category": "Product Category",
    "prod category": "Product Category", "product cat": "Product Category",
    "item category": "Product Category", "categ/ory": "Product Category",
}


def _norm_key(s: str) -> str:
    return re.sub(r"\s+", " ", str(s).strip().lower())


def _canonicalize_fixed_headers(df: pd.DataFrame) -> pd.DataFrame:
    """
    Rename header variants of the four fixed columns to their canonical names
    (Sl No / Item Code / Item Description / Product Category).

    - Date columns ("01/08 Op Stock", …) are left untouched.
    - A column already carrying the canonical name is never overwritten, and a
      rename that would collide with an existing canonical column is skipped, so
      duplicates are never introduced.
    """
    if df is None or df.empty:
        return df
    existing = set(df.columns)
    rename: dict[str, str] = {}
    for col in df.columns:
        if col in FIXED_COLS or _parse_col(col):
            continue  # already canonical, or a date sub-column
        canon = _FIXED_COL_SYNONYMS.get(_norm_key(col))
        if canon and canon not in existing:
            rename[col] = canon
            existing.discard(col)
            existing.add(canon)
    return df.rename(columns=rename) if rename else df


# ─── Credentials ──────────────────────────────────────────────────────────────

def _imap_creds() -> tuple[str, str]:
    e = os.getenv("EMAIL_SENDER", "").strip()
    p = os.getenv("EMAIL_PASSWORD", "").strip()
    if e and p:
        return e, p
    try:
        import streamlit as st
        try:
            return st.secrets["admin"]["EMAIL_SENDER"], st.secrets["admin"]["EMAIL_PASSWORD"]
        except Exception:
            return st.secrets["EMAIL_SENDER"], st.secrets["EMAIL_PASSWORD"]
    except Exception:
        return "", ""


def _email_recipients() -> list[str]:
    raw = (
        os.getenv("STOCK_34S_EMAIL_RECIPIENTS", "")
        or os.getenv("EMAIL_RECIPIENTS", "")
    )
    if raw:
        return [r.strip() for r in raw.split(",") if r.strip()]
    try:
        import streamlit as st
        try:
            raw = (
                st.secrets["admin"].get("STOCK_34S_EMAIL_RECIPIENTS")
                or st.secrets["admin"]["EMAIL_RECIPIENTS"]
            )
        except Exception:
            raw = (
                st.secrets.get("STOCK_34S_EMAIL_RECIPIENTS")
                or st.secrets["EMAIL_RECIPIENTS"]
            )
        return [r.strip() for r in raw.split(",") if r.strip()]
    except Exception:
        return []


# ─── Spreadsheet helpers ───────────────────────────────────────────────────────

def _get_spreadsheet():
    """Return authenticated gspread Spreadsheet object (always OPS spreadsheet)."""
    # First try to reuse the cached OPS client from sheets.py
    try:
        from services.sheets import _get_ops_spreadsheet
        return _get_ops_spreadsheet()
    except Exception:
        pass
    import json
    from google.oauth2.service_account import Credentials
    import gspread
    SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
    raw = os.getenv("GOOGLE_CREDENTIALS", "").strip()
    if raw:
        creds = Credentials.from_service_account_info(json.loads(raw), scopes=SCOPES)
    else:
        try:
            import streamlit as st
            creds = Credentials.from_service_account_info(st.secrets["google"], scopes=SCOPES)
        except Exception:
            raise RuntimeError("Google credentials not found")
    from services.sheet_config import OPS_SPREADSHEET_ID
    return gspread.authorize(creds).open_by_key(OPS_SPREADSHEET_ID)


def _worksheet_exists(sheet_name: str) -> bool:
    """
    Return True if the worksheet already exists in the OPS spreadsheet.

    This is a READ-ONLY check — unlike services.sheets.get_df (which calls
    _ensure_sheet and silently CREATES a missing tab), this never mutates the
    spreadsheet.  It guards load_month_df so that reading a month tab that
    doesn't exist yet (e.g. the previous month during carry-forward look-ups)
    can never leave behind a stray empty "34s Stock Register- <month>" sheet.
    """
    try:
        sh = _get_spreadsheet()
        sh.worksheet(sheet_name)
        return True
    except gspread.exceptions.WorksheetNotFound:
        return False
    except Exception:
        # On a transient/auth error, assume it exists so we don't mask real
        # data with a false "not found"; the subsequent read will surface it.
        return True


def _read_sheet_direct(sheet_name: str) -> pd.DataFrame:
    """
    Read a worksheet directly via gspread, bypassing the st.cache_data cache.
    Used inside catch-up loops where consecutive writes must be visible immediately.
    """
    try:
        sh = _get_spreadsheet()
        ws = sh.worksheet(sheet_name)
        data = ws.get_all_values()
        if not data:
            return pd.DataFrame()
        df = pd.DataFrame(data[1:], columns=data[0])
        df.columns = [str(c).strip() for c in df.columns]
        df = _canonicalize_fixed_headers(df)
        return df.replace("", pd.NA).dropna(how="all").fillna("").reset_index(drop=True)
    except Exception:
        return pd.DataFrame()


def _write_sheet_direct(sheet_name: str, df: pd.DataFrame) -> None:
    """Write a DataFrame to a worksheet directly via gspread."""
    from services.sheets import _serialize_for_sheets
    sh = _get_spreadsheet()
    try:
        ws = sh.worksheet(sheet_name)
    except gspread.exceptions.WorksheetNotFound:
        ws = sh.add_worksheet(title=sheet_name, rows=1200, cols=220)
    data = _serialize_for_sheets(df.fillna("").astype(str))
    ws.clear()
    ws.update("A1", data)


# ─── Month sheet discovery ─────────────────────────────────────────────────────

def get_available_months() -> list[tuple[int, int, str]]:
    """
    Return [(year, month, sheet_name), …] for all existing stock register tabs,
    sorted chronologically.
    """
    try:
        sh = _get_spreadsheet()
        result = []
        for ws in sh.worksheets():
            title = ws.title
            if not title.startswith(SHEET_PREFIX):
                continue
            suffix = title[len(SHEET_PREFIX):].strip()   # "May 2026"
            try:
                d = datetime.strptime(suffix, "%B %Y")
                result.append((d.year, d.month, title))
            except ValueError:
                pass
        result.sort()
        return result
    except Exception as e:
        print(f"[STOCK 34S] get_available_months error: {e}")
        return []


# ─── Load helpers ─────────────────────────────────────────────────────────────

def load_month_df(year: int, month: int, direct: bool = False) -> tuple[pd.DataFrame, str]:
    """
    Load the full horizontal DataFrame for a month tab.
    direct=True bypasses the st.cache_data cache (used in catch-up loops).
    """
    name = sheet_name_for(date(year, month, 1))
    if direct:
        df = _read_sheet_direct(name)
        if df.empty:
            return df, f"⚠️ Sheet '{name}' is empty or does not exist."
        return df, f"✅ {len(df)} items loaded (direct read)."

    # Guard: never let a read create the tab.  services.sheets.get_df routes
    # through _ensure_sheet, which auto-creates a missing worksheet — that would
    # spawn stray empty month tabs (e.g. when looking up the previous month for
    # Op Stock carry-forward).  Check existence first and bail out cleanly.
    if not _worksheet_exists(name):
        return pd.DataFrame(), f"⚠️ Sheet '{name}' does not exist yet."

    from services.sheets import get_df
    try:
        df = get_df(name)
    except Exception as e:
        return pd.DataFrame(), f"❌ Could not read '{name}': {e}"
    if df is None or df.empty:
        return pd.DataFrame(), f"⚠️ Sheet '{name}' is empty or does not exist yet."
    df.columns = [str(c).strip() for c in df.columns]
    df = _canonicalize_fixed_headers(df)
    df = df.replace("", pd.NA).dropna(how="all").fillna("").reset_index(drop=True)
    return df, f"✅ {len(df)} items loaded from '{name}'."


def load_stock_for_date(year: int, month: int, d: date) -> tuple[pd.DataFrame, str]:
    """
    Extract a flat display table (FIXED_COLS + DATE_SUB_COLS) for a single date.
    Returns (df, status_message).
    """
    df, status = load_month_df(year, month)
    if df.empty:
        return df, status

    available = dates_in_df(df, year, month)
    if d not in available:
        avail_str = ", ".join(x.strftime("%d/%m") for x in available) or "none"
        return pd.DataFrame(), (
            f"⚠️ No data for **{d.strftime('%d %b %Y')}** in this sheet.  \n"
            f"Dates with data: {avail_str}.  \n"
            "Add data manually or use **Force Update** to fetch automatically."
        )

    result = pd.DataFrame()
    for fc in FIXED_COLS:
        result[fc] = df.get(fc, pd.Series([""] * len(df))).values
    for sub in DATE_SUB_COLS:
        result[sub] = df.get(_col(d, sub), pd.Series([""] * len(df))).values

    return result.reset_index(drop=True), f"✅ {len(result)} items for {d.strftime('%d %b %Y')}."


def get_last_updated_date(year: int, month: int) -> date | None:
    """
    Return the most recent date whose Cl Stock column contains at least one
    non-empty value — i.e. the daily job actually ran for that date.

    Dates that only have column *headers* (added by Setup Sheet / seed_days)
    but no data values are NOT counted.  This prevents an empty seeded sheet
    from appearing to be "up to date".
    """
    df, _ = load_month_df(year, month)
    if df.empty:
        return None
    available = dates_in_df(df, year, month)
    if not available:
        return None

    dates_with_data: list[date] = []
    for d in available:
        cl_col = _col(d, "Cl Stock")
        if cl_col not in df.columns:
            continue
        # A cell is considered "written" if it is non-empty, non-NaN, and not
        # literally the string "nan" / "None" that pandas may leave behind.
        col_vals = df[cl_col].astype(str).str.strip()
        has_data = col_vals[~col_vals.isin(["", "nan", "None"])].any()
        if has_data:
            dates_with_data.append(d)

    return max(dates_with_data) if dates_with_data else None


# ─── Flat register: read/write engine ─────────────────────────────────────────

def _num(v) -> float:
    try:
        return float(str(v).replace(",", "").strip() or 0)
    except (ValueError, TypeError):
        return 0.0


def _fmt_int(x) -> int:
    try:
        return int(round(float(x)))
    except (ValueError, TypeError):
        return 0


def _read_raw_values(name: str):
    """Return (worksheet, all_values) for a tab, or (None, []) if it doesn't exist."""
    try:
        sh = _get_spreadsheet()
        ws = sh.worksheet(name)
    except gspread.exceptions.WorksheetNotFound:
        return None, []
    except Exception as e:
        print(f"[STOCK 34S] _read_raw_values open error for '{name}': {e}")
        return None, []
    try:
        return ws, ws.get_all_values()
    except Exception as e:
        print(f"[STOCK 34S] _read_raw_values read error for '{name}': {e}")
        return ws, []


def _find_header_row(values: list[list[str]]) -> int:
    """
    Return the 0-based index of the column-header row (the one carrying
    Item Code / Op Stock / Cl Stock labels), scanning past any title/banner row.
    Returns -1 if not found.
    """
    for i, row in enumerate(values[:15]):
        canon = {_FLAT_SYNONYMS.get(_norm_key(c)) for c in row}
        canon.discard(None)
        if "Item Code" in canon or {"Op Stock", "Cl Stock"} <= canon:
            return i
    return -1


def _flat_col_map(header: list[str]) -> dict[str, int]:
    """Map canonical column name → 0-based index using the flat synonym table."""
    colmap: dict[str, int] = {}
    for idx, cell in enumerate(header):
        canon = _FLAT_SYNONYMS.get(_norm_key(cell))
        if canon and canon not in colmap:
            colmap[canon] = idx
    return colmap


def _detect_item_code_col(values: list[list[str]], header_idx: int) -> int:
    """
    Return the 0-based index of the column that actually holds Godrej item codes,
    detected by content (majority of cells match _ITEM_CODE_RE).  Falls back to
    -1 when no column matches, so the caller can use the labelled column instead.
    """
    data = values[header_idx + 1: header_idx + 60]
    if not data:
        return -1
    ncol = max((len(r) for r in data), default=0)
    best_idx, best_hits = -1, 0
    for c in range(ncol):
        hits = sum(1 for row in data if c < len(row) and _ITEM_CODE_RE.search(str(row[c])))
        if hits > best_hits:
            best_idx, best_hits = c, hits
    return best_idx if best_hits > 0 else -1


def _parse_title_date(row: list[str]) -> tuple[date | None, int | None]:
    """Find a date in a title/banner row → (date, col_index) or (None, None)."""
    for j, cell in enumerate(row):
        s = str(cell).strip()
        if not s:
            continue
        for fmt in ("%d %B %Y", "%d %b %Y", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"):
            try:
                return datetime.strptime(s, fmt).date(), j
            except ValueError:
                continue
    return None, None


def is_flat_sheet(name: str) -> bool:
    """
    True if the tab uses the flat single-day register layout (title row + header
    row + one Op/In/Out/Cl block) rather than the horizontal date-pivot layout.
    """
    _, values = _read_raw_values(name)
    if not values:
        return False
    # Any date-prefixed column ("07/08 Op Stock") in the first rows ⇒ pivot layout
    for row in values[:3]:
        if any(_parse_col(str(cell).strip()) for cell in row):
            return False
    return _find_header_row(values) >= 0


def flat_as_of(name: str) -> date | None:
    """Return the register's current 'as of' date from its title banner, if any."""
    _, values = _read_raw_values(name)
    if not values:
        return None
    h = _find_header_row(values)
    if h < 1:
        return None
    d, _ = _parse_title_date(values[h - 1])
    return d


def load_flat_snapshot(name: str) -> tuple[pd.DataFrame, date | None, dict[str, str]]:
    """
    Load the flat register as a display DataFrame (exactly the user's own columns
    and header labels), plus the banner 'as of' date and a movement-column map
    {canonical_name: actual_header_label} so callers can locate Op/In/Out/Cl and
    the item-code column regardless of the sheet's exact wording.
    """
    _, values = _read_raw_values(name)
    if not values:
        return pd.DataFrame(), None, {}
    h = _find_header_row(values)
    if h < 0:
        return pd.DataFrame(), None, {}

    header = [str(c).strip() for c in values[h]]
    # De-duplicate blank / repeated header labels so pandas can build the frame
    counts: dict[str, int] = {}
    cols: list[str] = []
    for j, c in enumerate(header):
        base = c or f"Column {j + 1}"
        n = counts.get(base, 0)
        counts[base] = n + 1
        cols.append(base if n == 0 else f"{base} ({n + 1})")

    colmap_idx = _flat_col_map(header)
    code_idx = _detect_item_code_col(values, h)
    move_map: dict[str, str] = {}
    for canon, idx in colmap_idx.items():
        if canon in FLAT_MOVE_COLS and idx < len(cols):
            move_map[canon] = cols[idx]
    if code_idx >= 0 and code_idx < len(cols):
        move_map["Item Code"] = cols[code_idx]
    elif "Item Code" in colmap_idx:
        move_map["Item Code"] = cols[colmap_idx["Item Code"]]

    as_of, _ = _parse_title_date(values[h - 1]) if h >= 1 else (None, None)

    rows = []
    for r in range(h + 1, len(values)):
        raw = values[r]
        if not any(str(x).strip() for x in raw):
            continue
        rows.append([raw[j] if j < len(raw) else "" for j in range(len(cols))])

    df = pd.DataFrame(rows, columns=cols)
    return df, as_of, move_map


def run_flat_update(target_date: date | None = None) -> tuple[pd.DataFrame, str]:
    """
    Update the flat register IN PLACE for target_date (default: today IST).

    Non-destructive by design — only the four movement columns and the title
    date cell are rewritten (via targeted range updates); the sheet is never
    cleared, no columns or tabs are added.

      • New day  (banner date < target): Op Stock ← previous Cl Stock (carry
        forward); In/Out reset then filled from challan/delivery sources.
      • Same day (banner date == target): Op Stock kept as entered; In/Out taken
        from sources where found, otherwise the existing cell is preserved.
      • Cl Stock is always recomputed as Op + In − Out.
    """
    if target_date is None:
        target_date = datetime.now(IST).date()
    name = sheet_name_for(target_date)

    ws, values = _read_raw_values(name)
    if ws is None:
        return pd.DataFrame(), (
            f"❌ Sheet '{name}' not found. Create the month tab first — "
            "no new sheet was created."
        )
    if not values:
        return pd.DataFrame(), f"⚠️ Sheet '{name}' is empty."

    h = _find_header_row(values)
    if h < 0:
        return pd.DataFrame(), (
            f"❌ Could not find a header row (with Item Code / Op Stock / Cl Stock) "
            f"in '{name}'."
        )
    header = [str(c).strip() for c in values[h]]
    colmap = _flat_col_map(header)
    for req in ("Op Stock", "Cl Stock"):
        if req not in colmap:
            return pd.DataFrame(), f"❌ '{name}' has no '{req}' column. Headers: {header}."

    op_i, cl_i = colmap["Op Stock"], colmap["Cl Stock"]
    in_i, out_i = colmap.get("In Ward"), colmap.get("Out Ward")

    code_col = _detect_item_code_col(values, h)
    if code_col < 0:
        code_col = colmap.get("Item Code", -1)
    if code_col < 0:
        return pd.DataFrame(), (
            f"❌ Could not locate the item-code column in '{name}' "
            "(no column matches the Godrej item-code pattern)."
        )

    title_idx = h - 1 if h >= 1 else None
    as_of, date_col = (_parse_title_date(values[title_idx]) if title_idx is not None else (None, None))
    new_day = (as_of is None) or (as_of < target_date)

    # Movement sources (same as the pivot path)
    email_in = _fetch_inward_from_email(target_date)
    drive_in = _fetch_inward_from_drive(target_date)
    inward: dict[str, float] = {}
    for code, info in {**email_in, **drive_in}.items():
        code = code.upper()
        inward[code] = inward.get(code, 0.0) + float(info.get("qty", 0.0))
    outward = {k.upper(): v for k, v in _fetch_outward(target_date).items()}

    def cell(row: list[str], i: int | None) -> str:
        return row[i] if (i is not None and i < len(row)) else ""

    op_out, in_out, out_out, cl_out = [], [], [], []
    n_items = 0
    matched_in = matched_out = 0

    for r in range(h + 1, len(values)):
        row = values[r]
        code = str(cell(row, code_col)).strip().upper()
        cur_op, cur_cl = _num(cell(row, op_i)), _num(cell(row, cl_i))
        cur_in = _num(cell(row, in_i)) if in_i is not None else 0.0
        cur_out = _num(cell(row, out_i)) if out_i is not None else 0.0

        if not _ITEM_CODE_RE.search(code):
            # Not an item row (blank / subtotal / stray) — preserve existing cells
            op_out.append(cell(row, op_i))
            in_out.append(cell(row, in_i) if in_i is not None else "")
            out_out.append(cell(row, out_i) if out_i is not None else "")
            cl_out.append(cell(row, cl_i))
            continue

        n_items += 1
        op = cur_cl if new_day else cur_op
        if new_day:
            iw = inward.get(code, 0.0)
            ow = outward.get(code, 0.0)
        else:
            iw = inward.get(code, cur_in)
            ow = outward.get(code, cur_out)
        if code in inward:
            matched_in += 1
        if code in outward:
            matched_out += 1
        cl = op + iw - ow

        op_out.append(_fmt_int(op))
        in_out.append(_fmt_int(iw) if in_i is not None else "")
        out_out.append(_fmt_int(ow) if out_i is not None else "")
        cl_out.append(_fmt_int(cl))

    if n_items == 0:
        return pd.DataFrame(), (
            f"❌ No item rows found under the header in '{name}' "
            "(no cell matches the Godrej item-code pattern)."
        )

    # Targeted, non-destructive writes — only the movement columns + date cell
    from gspread.utils import rowcol_to_a1
    first = h + 2  # 1-based row number of the first data row

    def col_range(idx: int, vals: list) -> dict:
        a = rowcol_to_a1(first, idx + 1)
        b = rowcol_to_a1(first + len(vals) - 1, idx + 1)
        return {"range": f"{a}:{b}", "values": [[v] for v in vals]}

    updates = [col_range(op_i, op_out), col_range(cl_i, cl_out)]
    if in_i is not None:
        updates.append(col_range(in_i, in_out))
    if out_i is not None:
        updates.append(col_range(out_i, out_out))
    if title_idx is not None and date_col is not None:
        updates.append({
            "range": rowcol_to_a1(title_idx + 1, date_col + 1),
            "values": [[target_date.strftime("%d %B %Y")]],
        })

    try:
        ws.batch_update(updates, value_input_option="USER_ENTERED")
    except Exception as e:
        return pd.DataFrame(), f"❌ Write failed: {e}"

    mode = "carried forward from previous Cl Stock" if new_day else "kept as entered"
    status = (
        f"✅ '{name}' updated for {target_date.strftime('%d %b %Y')} — {n_items} items "
        f"(Op Stock {mode}). Inward matched: {matched_in} | Outward matched: {matched_out}."
    )
    print(f"[STOCK 34S] {status}")
    df, _, _ = load_flat_snapshot(name)
    return df, status


# ─── Sheet setup ──────────────────────────────────────────────────────────────

def ensure_month_sheet(year: int, month: int, seed_days: int = 0) -> str:
    """
    Create the month worksheet if it doesn't exist.

    - If the previous month's sheet exists, copies its item rows (FIXED_COLS)
      into the new sheet automatically so items don't need to be re-entered.
    - seed_days > 0: pre-adds column headers for the last `seed_days` days
      (up to today) so the user can fill in historical data manually.

    Returns a status message.
    """
    name = sheet_name_for(date(year, month, 1))
    try:
        sh = _get_spreadsheet()
    except Exception as e:
        return f"❌ Spreadsheet connection failed: {e}"

    # Check if it already exists — only catch WorksheetNotFound, not API/auth errors
    try:
        ws = sh.worksheet(name)
        # Sheet exists — check if it already has column headers
        try:
            existing_headers = ws.row_values(1)
        except Exception:
            existing_headers = []
        if existing_headers:
            return f"ℹ️ Sheet '{name}' already exists ({len(existing_headers)} columns)."
        # Sheet exists but is empty — fall through to write headers
    except gspread.exceptions.WorksheetNotFound:
        ws = sh.add_worksheet(title=name, rows=1200, cols=220)

    # Build header row
    headers = list(FIXED_COLS)
    today = datetime.now(IST).date()
    if seed_days > 0:
        for i in range(seed_days - 1, -1, -1):   # oldest → newest
            day_d = today - timedelta(days=i)
            if day_d.month == month and day_d.year == year:
                for sub in DATE_SUB_COLS:
                    headers.append(_col(day_d, sub))

    # Try to seed item rows from previous month
    prev_items_data = []
    first = date(year, month, 1)
    prev_last = first - timedelta(days=1)
    prev_df, _ = load_month_df(prev_last.year, prev_last.month)
    if not prev_df.empty:
        keep = [c for c in FIXED_COLS if c in prev_df.columns]
        items = prev_df[keep].copy()
        items = items[items.get("Item Code", pd.Series()).astype(str).str.strip() != ""]
        items = items.drop_duplicates(subset=["Item Code"])
        # Pad to match header length
        for col in FIXED_COLS:
            if col not in items.columns:
                items[col] = ""
        for h in headers:
            if h not in FIXED_COLS:
                items[h] = ""
        prev_items_data = items[headers].fillna("").astype(str).values.tolist()

    # Write header + item rows
    data = [headers] + prev_items_data
    ws.update("A1", data)

    item_note = f" + {len(prev_items_data)} items copied from {prev_last.strftime('%B %Y')}." if prev_items_data else " (no items yet — add item rows manually)."
    col_note  = f" + {seed_days} days of date columns ({len(headers)} columns total)" if seed_days else ""
    return f"✅ Sheet '{name}' created{col_note}{item_note}"


# ─── Previous Cl Stock look-up ────────────────────────────────────────────────

def _get_prev_cl_stock(
    target_date: date,
    df_override: pd.DataFrame | None = None,
) -> dict[str, float]:
    """
    Return {item_code_upper: cl_stock_float} from the most recent date
    STRICTLY BEFORE target_date that has Cl Stock data.

    Searches the current month first; if target_date is the 1st of the month
    (or current month has no earlier dates), falls back to the previous month tab.

    df_override: pass an already-loaded month DataFrame to skip a sheet read
                 (used inside catch-up loops for performance).
    """

    def _extract(df: pd.DataFrame, d: date) -> dict[str, float]:
        col = _col(d, "Cl Stock")
        if col not in df.columns:
            return {}
        result: dict[str, float] = {}
        for _, row in df.iterrows():
            code = str(row.get("Item Code", "")).strip().upper()
            if not code:
                continue
            try:
                result[code] = float(str(row.get(col, 0)).replace(",", "") or 0)
            except (ValueError, TypeError):
                result[code] = 0.0
        return result

    # 1. Current month
    if df_override is not None:
        df = df_override
    else:
        df, _ = load_month_df(target_date.year, target_date.month)

    if not df.empty:
        available = dates_in_df(df, target_date.year, target_date.month)
        past = [d for d in available if d < target_date]
        if past:
            prev = max(past)
            cl = _extract(df, prev)
            if cl:
                print(f"[STOCK 34S] Prev Cl from {prev} (current month)")
                return cl

    # 2. Previous month fallback (handles 1st of month or empty current month)
    first      = target_date.replace(day=1)
    prev_last  = first - timedelta(days=1)
    df2, _     = load_month_df(prev_last.year, prev_last.month)
    if not df2.empty:
        available2 = dates_in_df(df2, prev_last.year, prev_last.month)
        if available2:
            prev2 = max(available2)
            cl2 = _extract(df2, prev2)
            if cl2:
                print(f"[STOCK 34S] Prev Cl from {prev2} (previous month tab)")
                return cl2

    print("[STOCK 34S] No previous Cl Stock found — Op Stock = 0 (first run).")
    return {}


# ─── Master item list ──────────────────────────────────────────────────────────

def _get_master_items(year: int, month: int, df: pd.DataFrame | None = None) -> pd.DataFrame:
    """
    Return the item master (FIXED_COLS) from the month's sheet.
    Pass df directly to skip a re-read.

    NOTE: intentionally does NOT drop duplicates.  All rows with a non-empty
    Item Code are kept so that the row count always matches the full sheet.
    Removing duplicates here would cause a positional mismatch when writing
    new columns back to the sheet (e.g. 309 values for 319 rows).
    """
    if df is None:
        df, _ = load_month_df(year, month)
    if df.empty:
        return pd.DataFrame()
    keep = [c for c in FIXED_COLS if c in df.columns]
    if "Item Code" not in keep:
        return pd.DataFrame()
    items = df[keep].copy()
    items = items[items["Item Code"].astype(str).str.strip() != ""]
    return items.reset_index(drop=True)


def _count_items(df: pd.DataFrame) -> int:
    """Count rows in df that have a non-empty Item Code."""
    if df.empty or "Item Code" not in df.columns:
        return 0
    return int(df["Item Code"].astype(str).str.strip().ne("").sum())


# ─── PDF parsing (Delivery Challan) ───────────────────────────────────────────

def _parse_challan_pdf(pdf_bytes: bytes) -> dict:
    """
    Parse a Delivery Challan PDF.
    Returns {"warehouse_code": str, "challan_no": str,
             "items": {item_code: {"qty": float, "description": str}}}
    """
    result: dict = {"warehouse_code": "", "challan_no": "", "items": {}}
    text = ""
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            text = "\n".join(p.extract_text() or "" for p in pdf.pages)
    except Exception:
        try:
            import fitz
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            text = "\n".join(page.get_text() for page in doc)
        except Exception:
            return result

    wc = re.search(r"Warehouse\s*Code\s*[:\-]?\s*(ZB\w+)", text, re.IGNORECASE)
    if wc:
        result["warehouse_code"] = wc.group(1).strip()

    dc = re.search(r"Delivery\s+Challan\s+No\s*[:\-]?\s*([A-Z0-9]+)", text, re.IGNORECASE)
    if dc:
        result["challan_no"] = dc.group(1).strip()

    item_re = re.compile(r"(\d{8}[A-Z]{2}\d{5})")
    qty_re  = re.compile(r"\b(\d+(?:\.\d+)?)\s*(?:ECH|NOS|PCS|EACH|EA)?\b")
    for line in text.split("\n"):
        codes = item_re.findall(line)
        if not codes:
            continue
        qty_m = qty_re.search(line)
        qty   = float(qty_m.group(1)) if qty_m else 1.0
        for code in codes:
            if code in result["items"]:
                result["items"][code]["qty"] += qty
            else:
                result["items"][code] = {"qty": qty, "description": line.strip()[:80]}
    return result


# ─── Inward — email ───────────────────────────────────────────────────────────

def _fetch_inward_from_email(target_date: date) -> dict[str, dict]:
    """Search for 'Delivery Challan Information' emails on target_date and parse PDFs."""
    import imaplib
    import email as _email_lib

    email_addr, password = _imap_creds()
    if not email_addr or not password:
        print("[STOCK 34S] IMAP credentials not configured — skipping email inward.")
        return {}
    try:
        mail = imaplib.IMAP4_SSL("imap.gmail.com")
        mail.login(email_addr, password)
        mail.select("inbox")
    except Exception as e:
        print(f"[STOCK 34S] IMAP login failed: {e}")
        return {}

    since = target_date.strftime("%d-%b-%Y")
    until = (target_date + timedelta(days=1)).strftime("%d-%b-%Y")
    query = f'(SUBJECT "{CHALLAN_SUBJECT}" SINCE {since} BEFORE {until})'
    try:
        _, data = mail.search(None, query)
    except Exception as e:
        mail.logout()
        print(f"[STOCK 34S] IMAP search failed: {e}")
        return {}

    ids = data[0].split() if data and data[0] else []
    combined: dict[str, dict] = {}

    for eid in ids:
        try:
            _, msg_data = mail.fetch(eid, "(RFC822)")
            msg = _email_lib.message_from_bytes(msg_data[0][1])
        except Exception:
            continue
        for part in msg.walk():
            if part.get_content_type() != "application/pdf":
                continue
            pdf_bytes = part.get_payload(decode=True)
            if not pdf_bytes:
                continue
            parsed = _parse_challan_pdf(pdf_bytes)
            if WAREHOUSE_CODE not in parsed["warehouse_code"].upper():
                continue
            for code, info in parsed["items"].items():
                if code in combined:
                    combined[code]["qty"] += info["qty"]
                else:
                    combined[code] = {
                        "qty": info["qty"],
                        "description": info["description"],
                        "challan_no": parsed["challan_no"],
                    }
    mail.logout()
    print(f"[STOCK 34S] Email inward: {len(combined)} items for {target_date}")
    return combined


# ─── Inward — Google Drive ────────────────────────────────────────────────────

def _fetch_inward_from_drive(target_date: date) -> dict[str, dict]:
    """
    Read invoice PDFs from Google Drive for target_date.
    Reuses GOOGLE_DRIVE_INVOICES_FOLDER_ID and helpers from
    email_sender_delivery_schedule.py.
    Folder structure: <root> / "{N} {Month}-{YYYY}" / "for {DD}.{MM}.{YYYY}" / *.pdf
    """
    try:
        from services.email_sender_delivery_schedule import (
            _get_drive_folder_id, _get_drive_service,
            _list_drive_folders, _list_drive_pdfs, _download_drive_file,
        )
    except Exception as e:
        print(f"[STOCK 34S] Drive import error: {e}")
        return {}

    root_id = _get_drive_folder_id()
    if not root_id:
        print("[STOCK 34S] GOOGLE_DRIVE_INVOICES_FOLDER_ID not set — skipping Drive inward.")
        return {}
    try:
        svc = _get_drive_service()
    except Exception as e:
        print(f"[STOCK 34S] Drive service error: {e}")
        return {}

    fin_month = (target_date.month - 4) % 12 + 1
    mf_name   = f"{fin_month} {target_date.strftime('%B')}-{target_date.year}"
    df_name   = f"for {target_date.day:02d}.{target_date.month:02d}.{target_date.year}"

    def _find_folder(parent_id, name):
        for f in _list_drive_folders(svc, parent_id):
            if f.get("name") == name:
                return f["id"]
        return None

    mf_id = _find_folder(root_id, mf_name)
    if not mf_id:
        print(f"[STOCK 34S] Drive: '{mf_name}' not found.")
        return {}
    df_id = _find_folder(mf_id, df_name)
    if not df_id:
        print(f"[STOCK 34S] Drive: '{df_name}' not found.")
        return {}

    pdfs = _list_drive_pdfs(svc, df_id)
    combined: dict[str, dict] = {}
    for f in pdfs:
        content = _download_drive_file(svc, f["id"])
        if not content:
            continue
        try:
            parsed = _parse_challan_pdf(content)
        except Exception as e:
            print(f"[STOCK 34S] Drive parse error '{f['name']}': {e}")
            continue
        if WAREHOUSE_CODE not in parsed["warehouse_code"].upper():
            continue
        for code, info in parsed["items"].items():
            if code in combined:
                combined[code]["qty"] += info["qty"]
            else:
                combined[code] = {
                    "qty": info["qty"],
                    "description": info["description"],
                    "challan_no": parsed["challan_no"],
                }
    print(f"[STOCK 34S] Drive inward: {len(combined)} items for {target_date}")
    return combined


# ─── Outward from sheets ───────────────────────────────────────────────────────

def _fetch_outward(target_date: date) -> dict[str, float]:
    """
    Sum outward quantities from DELIVERY_SHEET + RETURN_SHEET for target_date.
    Returns {item_code_upper: total_qty}.
    """
    from services.sheets import get_df
    combined: dict[str, float] = {}
    target_str = target_date.strftime("%d/%m/%Y")
    # Also try without leading zeros in case the sheet uses that format
    alt_str    = f"{target_date.day}/{target_date.month}/{target_date.year}"

    for sheet_name in [DELIVERY_SHEET, RETURN_SHEET]:
        try:
            df = get_df(sheet_name)
        except Exception:
            continue
        if df is None or df.empty:
            continue
        df.columns = [str(c).strip().upper() for c in df.columns]
        if "DATE" not in df.columns:
            continue
        date_col = df["DATE"].astype(str).str.strip()
        mask = date_col.isin([target_str, alt_str])
        for _, row in df[mask].iterrows():
            code = str(row.get("ITEM CODE", "")).strip().upper()
            if not code:
                continue
            try:
                qty = float(str(row.get("QUANTITY", 0)).replace(",", "") or 0)
            except (ValueError, TypeError):
                qty = 0.0
            combined[code] = combined.get(code, 0.0) + qty

    print(f"[STOCK 34S] Outward: {len(combined)} items for {target_date}")
    return combined


# ─── Core: build columns for one date ─────────────────────────────────────────

def _build_date_columns(
    df: pd.DataFrame,
    target_date: date,
    prev_cl: dict[str, float],
    inward: dict[str, dict],
    outward: dict[str, float],
) -> dict[str, list]:
    """
    Build 5 new column lists (Op Stock / In Ward / Out Ward / Cl Stock / DC No)
    aligned row-for-row with the FULL sheet DataFrame `df`.

    Rows with an empty Item Code (blank rows, header rows, etc.) are preserved
    as empty strings so the output length always equals len(df).  This prevents
    the positional mismatch that caused some items to receive blank values.

    Logic per item row:
      Op Stock  = most recent Cl Stock for that item (from prev_cl lookup)
      In Ward   = quantity received (from email/Drive challan PDFs), else 0
      Out Ward  = quantity dispatched (from delivery/return sheets), else 0
      Cl Stock  = Op Stock + In Ward − Out Ward
      DC No     = Delivery Challan number if inward exists, else ""
    """
    op_vals, in_vals, out_vals, cl_vals, dc_vals = [], [], [], [], []

    for _, row in df.iterrows():
        code = str(row.get("Item Code", "")).strip().upper()
        if not code:
            # Blank / header / subtotal row — leave all 5 cells empty
            op_vals.append("")
            in_vals.append("")
            out_vals.append("")
            cl_vals.append("")
            dc_vals.append("")
            continue

        op = prev_cl.get(code, 0.0)
        iw = inward.get(code, {}).get("qty", 0.0)
        ow = outward.get(code, 0.0)
        cl = op + iw - ow
        dc = inward.get(code, {}).get("challan_no", "")

        op_vals.append(int(round(op)))
        in_vals.append(int(round(iw)) if iw else 0)
        out_vals.append(int(round(ow)) if ow else 0)
        cl_vals.append(int(round(cl)))
        dc_vals.append(dc)

    return {
        _col(target_date, "Op Stock"):  op_vals,
        _col(target_date, "In Ward"):   in_vals,
        _col(target_date, "Out Ward"):  out_vals,
        _col(target_date, "Cl Stock"):  cl_vals,
        _col(target_date, "DC No"):     dc_vals,
    }


# ─── Daily update (single date) ───────────────────────────────────────────────

def run_daily_update(
    target_date: date | None = None,
    df_in: pd.DataFrame | None = None,
    direct: bool = False,
) -> tuple[pd.DataFrame, str]:
    """
    Update the horizontal month sheet for target_date (default: today IST).
    Idempotent — replaces today's columns if already present.

    df_in   : pass an existing in-memory DataFrame to avoid a re-read
              (used inside run_update_range for performance).
    direct  : if True, use direct gspread reads/writes (bypasses cache).

    Returns (flat_df_for_date, status_message).
    """
    if target_date is None:
        target_date = datetime.now(IST).date()

    year, month = target_date.year, target_date.month
    name = sheet_name_for(target_date)
    print(f"[STOCK 34S] === Daily update for {target_date} ({name}) ===")

    # Flat single-day register (the live OPS layout) → update in place, never
    # touching column structure or creating anything.
    if df_in is None and is_flat_sheet(name):
        return run_flat_update(target_date)

    # 1. Ensure month sheet exists (creates it + copies previous month items if new)
    try:
        setup_msg = ensure_month_sheet(year, month, seed_days=0)
    except Exception as setup_err:
        return pd.DataFrame(), f"❌ Sheet setup failed: {setup_err}"
    print(f"[STOCK 34S] {setup_msg}")

    # 2. Load the month DataFrame
    if df_in is not None:
        df = df_in
    elif direct:
        df = _read_sheet_direct(name)
        if df.empty:
            return pd.DataFrame(), f"⚠️ Sheet '{name}' is empty — add item rows first."
    else:
        df, load_msg = load_month_df(year, month)
        if df.empty:
            return pd.DataFrame(), load_msg

    # 3. Validate item count (must have at least one row with Item Code)
    item_count = _count_items(df)
    if item_count == 0:
        cols_seen = ", ".join(str(c) for c in df.columns[:8]) or "(none)"
        has_code_col = "Item Code" in df.columns
        hint = (
            "The 'Item Code' column exists but has no filled rows — add item rows."
            if has_code_col else
            f"No 'Item Code' column was found. Columns detected: [{cols_seen}]. "
            "Rename the item-code column to 'Item Code' (casing/spacing and common "
            "variants like 'Material Code' are handled automatically)."
        )
        return pd.DataFrame(), f"❌ No items found in '{name}'. {hint}"

    # 4. Previous closing stock
    prev_cl = _get_prev_cl_stock(target_date, df_override=df)

    # 5. Inward — email challan PDFs + Google Drive invoice PDFs
    email_in = _fetch_inward_from_email(target_date)
    drive_in = _fetch_inward_from_drive(target_date)
    inward: dict[str, dict] = {}
    for code, info in {**email_in, **drive_in}.items():
        code = code.upper()
        if code in inward:
            inward[code]["qty"] += info.get("qty", 0.0)
        else:
            inward[code] = dict(info)

    # 6. Outward — 34S PHYSICAL DELIVERY CHALLAN + 34S RETURN RPL sheets
    outward = {k.upper(): v for k, v in _fetch_outward(target_date).items()}

    print(f"[STOCK 34S] Inward items found : {len(inward)}  (email:{len(email_in)} + drive:{len(drive_in)})")
    print(f"[STOCK 34S] Outward items found: {len(outward)}")

    # 7. Build 5 new columns — iterates over ALL rows in df (no filtering/dedup)
    new_cols = _build_date_columns(df, target_date, prev_cl, inward, outward)

    # 8. Remove existing columns for target_date (idempotent)
    tag  = _col_tag(target_date)
    drop = [c for c in df.columns if c.startswith(tag + " ")]
    df   = df.drop(columns=drop, errors="ignore")

    # 9. Assign new columns — lengths are guaranteed to match len(df)
    for col_name, values in new_cols.items():
        df[col_name] = values

    # 10. Write back
    try:
        if direct:
            _write_sheet_direct(name, df)
        else:
            from services.sheets import write_df
            write_df(name, df.fillna("").astype(str))
        status = (
            f"✅ '{name}' updated for {target_date} — {item_count} items.  "
            f"Inward: {len(inward)} item(s) | Outward: {len(outward)} item(s)."
        )
    except Exception as e:
        status = f"❌ Write failed: {e}"

    print(f"[STOCK 34S] {status}")
    flat, _ = load_stock_for_date(year, month, target_date)
    return flat, status


# ─── Catch-up: update a range of dates ────────────────────────────────────────

def run_update_range(start_date: date, end_date: date) -> tuple[list[str], str]:
    """
    Update every date from start_date to end_date inclusive.
    Maintains the DataFrame in memory between days (no cache issue) and writes
    to the sheet after each day.  Handles month crossings automatically.

    Returns (list_of_per_day_status_lines, summary_message).
    """
    # Flat single-day register keeps no per-day history — a "range" update just
    # refreshes today's in-place snapshot once.
    if is_flat_sheet(sheet_name_for(end_date)):
        _, status = run_flat_update(end_date)
        line = f"{end_date.strftime('%d/%m/%Y')}: {status}"
        return [line], status

    results: list[str] = []
    d       = start_date
    # Cache the current month DataFrame across days for performance
    cur_year  = d.year
    cur_month = d.month
    df_cache, _ = load_month_df(cur_year, cur_month, direct=True)
    name_cache  = sheet_name_for(date(cur_year, cur_month, 1))

    while d <= end_date:
        # Switch to new month if we've crossed a boundary
        if d.year != cur_year or d.month != cur_month:
            # Write current month before switching
            if not df_cache.empty:
                try:
                    _write_sheet_direct(name_cache, df_cache)
                except Exception as we:
                    results.append(f"  ⚠️ Write error for {cur_year}/{cur_month}: {we}")
            cur_year, cur_month = d.year, d.month
            ensure_month_sheet(cur_year, cur_month, seed_days=0)
            df_cache, _ = load_month_df(cur_year, cur_month, direct=True)
            name_cache  = sheet_name_for(date(cur_year, cur_month, 1))

        year, month = d.year, d.month
        name = sheet_name_for(d)

        item_count = _count_items(df_cache)
        if item_count == 0:
            cols_seen = ", ".join(str(c) for c in df_cache.columns[:8]) or "(none)"
            results.append(
                f"{d.strftime('%d/%m/%Y')}: ❌ No items with 'Item Code' found in sheet. "
                f"Columns detected: [{cols_seen}]."
            )
            d += timedelta(days=1)
            continue

        prev_cl  = _get_prev_cl_stock(d, df_override=df_cache)
        email_in = _fetch_inward_from_email(d)
        drive_in = _fetch_inward_from_drive(d)
        inward: dict[str, dict] = {}
        for code, info in {**email_in, **drive_in}.items():
            code = code.upper()
            if code in inward:
                inward[code]["qty"] += info.get("qty", 0.0)
            else:
                inward[code] = dict(info)
        outward = {k.upper(): v for k, v in _fetch_outward(d).items()}

        # Build columns over full df — no item filtering, no positional mismatch
        new_cols = _build_date_columns(df_cache, d, prev_cl, inward, outward)

        tag  = _col_tag(d)
        drop = [c for c in df_cache.columns if c.startswith(tag + " ")]
        df_cache = df_cache.drop(columns=drop, errors="ignore")
        for col_name, values in new_cols.items():
            df_cache[col_name] = values   # lengths guaranteed equal

        # Write after each day so the sheet is always up-to-date
        try:
            _write_sheet_direct(name, df_cache)
            status = (
                f"✅ {item_count} items | "
                f"Inward: {len(inward)} | Outward: {len(outward)}"
            )
        except Exception as e:
            status = f"❌ Write failed: {e}"

        results.append(f"{d.strftime('%d/%m/%Y')}: {status}")
        d += timedelta(days=1)

    total_ok  = sum(1 for r in results if "✅" in r)
    total_err = sum(1 for r in results if "❌" in r)
    summary = f"Catch-up complete: {total_ok} day(s) updated, {total_err} error(s)."
    return results, summary


# ─── Monthly email helpers ────────────────────────────────────────────────────

def _build_monthly_excel(year: int, month: int) -> bytes:
    """
    Export the full horizontal sheet for a month as a styled Excel file.
    Returns the raw .xlsx bytes.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    df, _ = load_month_df(year, month)
    wb    = openpyxl.Workbook()
    ws    = wb.active
    ws.title = sheet_name_for(date(year, month, 1))[len(SHEET_PREFIX):]

    if df.empty:
        ws.append(["No data found."])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # Header
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    date_fills  = [PatternFill("solid", fgColor=c) for c in
                   ["2E75B6", "2F5496", "1F3864", "203864", "17375E"]]

    headers = list(df.columns)
    ws.append(headers)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci)
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        parsed = _parse_col(h)
        if parsed:
            sub_idx = DATE_SUB_COLS.index(parsed[2]) if parsed[2] in DATE_SUB_COLS else 0
            cell.fill = date_fills[sub_idx % len(date_fills)]
        else:
            cell.fill = header_fill

    # Data rows
    for _, row in df.iterrows():
        ws.append(row.tolist())

    # Column widths
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = (
            22 if h in FIXED_COLS else 12
        )
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "E2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_stock_html_table(flat: pd.DataFrame) -> str:
    """
    Build a styled HTML table from a flat daily stock DataFrame
    (columns: FIXED_COLS + DATE_SUB_COLS).
    Shared by both the monthly and daily email functions.
    """
    th_style = "padding:8px 10px;background:#1F4E79;color:#fff;border:1px solid #ccc;white-space:nowrap;"
    td_style = "padding:6px 10px;border:1px solid #ddd;"
    cols     = [c for c in FIXED_COLS + DATE_SUB_COLS if c in flat.columns]
    ths      = "".join(f"<th style='{th_style}'>{c}</th>" for c in cols)
    rows_html = ""
    for _, r in flat.iterrows():
        # Highlight zero-stock rows
        try:
            cl_val = float(str(r.get("Cl Stock", 1)).replace(",", "") or 1)
            row_bg = "background:#FFEBEE;" if cl_val == 0 else ""
        except (ValueError, TypeError):
            row_bg = ""
        cells = "".join(
            f"<td style='{td_style}{row_bg}'>{r.get(c, '')}</td>" for c in cols
        )
        rows_html += f"<tr>{cells}</tr>"
    return (
        f"<table style='border-collapse:collapse;font-family:Arial,sans-serif;"
        f"font-size:12px;width:100%'>"
        f"<thead><tr>{ths}</tr></thead>"
        f"<tbody>{rows_html}</tbody></table>"
    )


def _build_generic_html_table(df: pd.DataFrame, cl_col: str | None = None) -> str:
    """
    Build a styled HTML table from any DataFrame, keeping its own columns/labels
    (used for the flat register, whose headers differ from FIXED/DATE_SUB_COLS).
    Rows whose `cl_col` value is 0 are highlighted.
    """
    th_style = "padding:8px 10px;background:#1F4E79;color:#fff;border:1px solid #ccc;white-space:nowrap;"
    td_style = "padding:6px 10px;border:1px solid #ddd;"
    cols     = list(df.columns)
    ths      = "".join(f"<th style='{th_style}'>{c}</th>" for c in cols)
    rows_html = ""
    for _, r in df.iterrows():
        row_bg = ""
        if cl_col and cl_col in df.columns:
            try:
                if float(str(r.get(cl_col, 1)).replace(",", "") or 1) == 0:
                    row_bg = "background:#FFEBEE;"
            except (ValueError, TypeError):
                row_bg = ""
        cells = "".join(f"<td style='{td_style}{row_bg}'>{r.get(c, '')}</td>" for c in cols)
        rows_html += f"<tr>{cells}</tr>"
    return (
        f"<table style='border-collapse:collapse;font-family:Arial,sans-serif;"
        f"font-size:12px;width:100%'>"
        f"<thead><tr>{ths}</tr></thead><tbody>{rows_html}</tbody></table>"
    )


def _build_flat_excel(name: str, df: pd.DataFrame) -> bytes:
    """Export the flat register snapshot as a simple styled .xlsx."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = name[len(SHEET_PREFIX):] if name.startswith(SHEET_PREFIX) else "Stock"
    if df.empty:
        ws.append(["No data found."])
    else:
        ws.append(list(df.columns))
        fill = PatternFill("solid", fgColor="1F4E79")
        font = Font(color="FFFFFF", bold=True, size=10)
        for ci, _ in enumerate(df.columns, 1):
            c = ws.cell(1, ci)
            c.fill, c.font = fill, font
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        for _, row in df.iterrows():
            ws.append(row.tolist())
        for ci, h in enumerate(df.columns, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 16
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_last_day_html_table(year: int, month: int) -> tuple[str, date | None]:
    """Return (html_table_string, last_date) for the last recorded day of the month."""
    df, _ = load_month_df(year, month)
    if df.empty:
        return "<p>No data found for this month.</p>", None
    available = dates_in_df(df, year, month)
    if not available:
        return "<p>No date columns found in the sheet.</p>", None
    last_day = max(available)
    flat, _  = load_stock_for_date(year, month, last_day)
    if flat.empty:
        return f"<p>No data for {last_day}.</p>", last_day
    return _build_stock_html_table(flat), last_day


def send_monthly_stock_email(year: int, month: int, archive: bool = True) -> dict:
    """
    Send the monthly stock email:
      - Subject : "Monthly 34S Stock Details- {Month} {Year}"
      - Body    : last recorded day's stock as an HTML table
      - Attach  : full month's data as a styled Excel file

    archive=True  (default, used by the scheduled job) also renames the sheet to
                  "ARCHIVED 34S Stock {Month} {Year}" after sending.
    archive=False (used by the manual UI button) sends the email only.
    """
    import smtplib
    from email.message import EmailMessage

    month_label = date(year, month, 1).strftime("%B %Y")
    subject     = f"Monthly 34S Stock Details- {month_label}"
    name        = sheet_name_for(date(year, month, 1))

    # Flat register → build report from the in-place snapshot.
    if is_flat_sheet(name):
        flat_df, as_of, move_map = load_flat_snapshot(name)
        if flat_df.empty:
            return {"sent": False, "error": f"No data found in '{name}'."}
        html_table   = _build_generic_html_table(flat_df, move_map.get("Cl Stock"))
        last_day     = as_of
        excel_bytes  = _build_flat_excel(name, flat_df)
    else:
        html_table, last_day = build_last_day_html_table(year, month)
        excel_bytes  = _build_monthly_excel(year, month)
    last_day_str = last_day.strftime("%d %b %Y") if last_day else "—"

    recipients = _email_recipients()
    if not recipients:
        return {"sent": False, "error": "No email recipients configured."}

    email_addr, password = _imap_creds()
    if not email_addr or not password:
        return {"sent": False, "error": "Email credentials not configured."}

    html_body = (
        f"<html><body style='font-family:Arial,sans-serif;color:#222'>"
        f"<div style='background:#1F4E79;padding:16px 24px;border-radius:6px 6px 0 0'>"
        f"<h2 style='color:#fff;margin:0'>📦 34S Physical Stock Register — {month_label}</h2>"
        f"<p style='color:#c5cae9;margin:6px 0 0;font-size:13px'>"
        f"Last recorded day: {last_day_str}</p></div>"
        f"<div style='padding:20px 24px'>"
        f"<p>Please find the month-end stock summary below and the full month's data attached.</p>"
        f"{html_table}"
        f"</div>"
        f"<div style='padding:10px 24px;background:#eee;font-size:11px;color:#888;"
        f"border-radius:0 0 6px 6px'>Automated monthly report · 34S Interiors CRM</div>"
        f"</body></html>"
    )

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"]    = email_addr
    msg["To"]      = ", ".join(recipients)
    msg.set_content("Please view this email in HTML format.")
    msg.add_alternative(html_body, subtype="html")
    msg.add_attachment(
        excel_bytes,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"34S_Stock_{month_label.replace(' ', '_')}.xlsx",
    )

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(email_addr, password)
            server.send_message(msg)
    except Exception as e:
        return {"sent": False, "error": str(e)}

    if archive:
        try:
            sh = _get_spreadsheet()
            old_name = sheet_name_for(date(year, month, 1))
            ws = sh.worksheet(old_name)
            ws.update_title(f"ARCHIVED 34S Stock {month_label}")
            print(f"[STOCK 34S] Sheet archived as 'ARCHIVED 34S Stock {month_label}'.")
        except Exception as e:
            print(f"[STOCK 34S] Archive rename failed (non-fatal): {e}")

    try:
        from services.sheets import append_email_log
        append_email_log(
            f"Monthly 34S Stock Email ({month_label})",
            0, recipients, "success",
        )
    except Exception:
        pass

    return {"sent": True, "subject": subject, "recipients": recipients}


def send_daily_stock_email(target_date: date) -> dict:
    """
    Send a daily stock snapshot email for target_date:
      - Subject : "34S Stock Report — {DD Month YYYY}"
      - Body    : that day's full stock table as HTML (no attachment)
      - Recipients: same EMAIL_RECIPIENTS secret used for all other CRM emails

    This is triggered manually from the dashboard; it is NOT scheduled.
    """
    import smtplib
    from email.message import EmailMessage

    year, month = target_date.year, target_date.month
    date_label  = target_date.strftime("%d %B %Y")
    day_name    = target_date.strftime("%A")
    subject     = f"34S Stock Report — {date_label}"
    name        = sheet_name_for(target_date)

    # Flat register → snapshot table with the sheet's own columns.
    if is_flat_sheet(name):
        flat, _as_of, move_map = load_flat_snapshot(name)
        cl_col, in_col, out_col = (
            move_map.get("Cl Stock"), move_map.get("In Ward"), move_map.get("Out Ward"),
        )
    else:
        flat, load_status = load_stock_for_date(year, month, target_date)
        cl_col, in_col, out_col = "Cl Stock", "In Ward", "Out Ward"
        move_map = None
    if flat.empty:
        return {
            "sent": False,
            "error": (
                f"No stock data found for {date_label}. "
                "Run ⚡ Update Sheet first to populate this date."
            ),
        }

    recipients = _email_recipients()
    if not recipients:
        return {"sent": False, "error": "No email recipients configured (EMAIL_RECIPIENTS secret)."}

    email_addr, password = _imap_creds()
    if not email_addr or not password:
        return {"sent": False, "error": "Email credentials not configured."}

    html_table = (
        _build_generic_html_table(flat, cl_col) if move_map is not None
        else _build_stock_html_table(flat)
    )

    # Summary line
    try:
        cl_total  = int(pd.to_numeric(flat.get(cl_col,  pd.Series()), errors="coerce").fillna(0).sum())
        in_total  = int(pd.to_numeric(flat.get(in_col,   pd.Series()), errors="coerce").fillna(0).sum())
        out_total = int(pd.to_numeric(flat.get(out_col,  pd.Series()), errors="coerce").fillna(0).sum())
        summary_line = (
            f"Total items: <b>{to_indian_number_string(len(flat), 0)}</b> &nbsp;|&nbsp; "
            f"Cl Stock: <b>{to_indian_number_string(cl_total, 0)}</b> &nbsp;|&nbsp; "
            f"In Ward: <b>{to_indian_number_string(in_total, 0)}</b> &nbsp;|&nbsp; "
            f"Out Ward: <b>{to_indian_number_string(out_total, 0)}</b>"
        )
    except Exception:
        summary_line = f"Total items: <b>{to_indian_number_string(len(flat), 0)}</b>"

    html_body = (
        f"<html><body style='font-family:Arial,sans-serif;color:#222'>"
        f"<div style='background:#1F4E79;padding:16px 24px;border-radius:6px 6px 0 0'>"
        f"<h2 style='color:#fff;margin:0'>📦 34S Stock Report — {date_label} ({day_name})</h2>"
        f"</div>"
        f"<div style='padding:16px 24px 8px'>"
        f"<p style='font-size:13px;color:#555;margin:0 0 12px'>{summary_line}</p>"
        f"{html_table}"
        f"</div>"
        f"<div style='padding:10px 24px;background:#eee;font-size:11px;color:#888;"
        f"border-radius:0 0 6px 6px'>Daily snapshot · 34S Interiors CRM</div>"
        f"</body></html>"
    )

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"]    = email_addr
    msg["To"]      = ", ".join(recipients)
    msg.set_content("Please view this email in HTML format.")
    msg.add_alternative(html_body, subtype="html")

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(email_addr, password)
            server.send_message(msg)
    except Exception as e:
        return {"sent": False, "error": str(e)}

    try:
        from services.sheets import append_email_log
        append_email_log(
            f"Daily 34S Stock Email ({date_label})",
            len(flat), recipients, "success",
        )
    except Exception:
        pass

    return {"sent": True, "subject": subject, "recipients": recipients}
